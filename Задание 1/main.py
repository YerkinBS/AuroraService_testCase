# начало создания скрипта
import os
import sys
import logging
from openpyxl import load_workbook
from docx2pdf import convert
from tkinter import Tk, filedialog, messagebox
from fpdf import FPDF

# настройка логирования в файл logs.txt
logging.basicConfig(filename="logs.txt", level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logging.info("\n" + "="*50 + "\nЗапуск выполнения нового скрипта\n" + "="*50)

# функция для избежания перезаписи файлов, если они уже существуют в папке pdfs, то создаются новые с индексом
def Check_If_Exists(file_path):
    logging.info(f"Проверка существования файла {file_path}")
    cnt = 1
    while os.path.exists(f"{file_path} ({cnt}).pdf"):
        cnt += 1
        logging.info(f"Файл '{file_path} ({cnt - 1}).pdf' уже существует")
    logging.info(f"Файл '{file_path} ({cnt}).pdf' не существует. Создание нового файла")
    return f"{file_path} ({cnt}).pdf"


# функция для конвертации файла Excel в PDF с помощью библиотеки reportlab и openpyxl
# важно! при выборе файла Excel необходимо выбирать файл с расширением .xlsx, ибо .xls не поддерживается
def xlsx_to_pdf(file_path):
    try:
        wb = load_workbook(file_path)
        worksheet = wb.active

        # cоздание директории для PDF файлов, если она не существует
        pdf_dir = os.path.join(os.path.dirname(__file__), "pdfs")
        os.makedirs(pdf_dir, exist_ok=True)

        # создание файла PDF в папке pdfs
        # проверка существования файла, если файл с таким именем уже существует, то создается новый с индексом
        pdf_output_path = Check_If_Exists(os.path.join(pdf_dir, os.path.splitext(os.path.basename(file_path))[0]))

        logging.info(f"Создание файла {pdf_output_path}")

        # инициализация PDF
        pdf = FPDF(orientation='L')
        # pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        # определение параметров страницы и размеров ячеек
        page_width = pdf.w - 2 * pdf.l_margin
        page_height = pdf.h - 2 * pdf.t_margin
        cell_padding = 2  # добавим небольшие отступы между содержимым ячеек и границами

        # определение максимальных размеров ячеек
        max_cell_widths = [0] * worksheet.max_column
        for row in worksheet.iter_rows():
            for i, cell in enumerate(row):
                cell_value = str(cell.value)
                cell_width = pdf.get_string_width(cell_value) + 2 * cell_padding
                max_cell_widths[i] = max(max_cell_widths[i], cell_width)

        # создание PDF с учетом максимальных размеров ячеек
        current_y = pdf.t_margin
        for row in worksheet.iter_rows():
            for i, cell in enumerate(row):
                cell_value = str(cell.value)
                if pdf.get_y() + 10 > pdf.page_break_trigger:
                    pdf.add_page()
                    current_y = pdf.t_margin
                pdf.cell(max_cell_widths[i], 10, cell_value, border=1)
                current_y += 10
            pdf.ln()

        # сохранение PDF
        pdf.output(pdf_output_path)
        logging.info(f"Файл {pdf_output_path} успешно создан")
    except Exception as e:  # обработка исключений
        logging.error(f"Ошибка при загрузке файла Excel: {e}")
        messagebox.showerror("Ошибка", f"Ошибка при загрузке файла Excel: {e}")
        sys.exit(1)


# функция для конвертации файла Word в PDF с помощью библиотеки docx2pdf
# важно! при выборе файла Word необходимо выбирать файл с расширением .docx, ибо .doc не поддерживается
def convert_docx_to_pdf(file_path):
    logging.info(f"Создание файла {file_path}")
    try:
        # создание файла PDF в папке pdfs
        script_dir = f"{os.path.dirname(__file__)}/pdfs"
        pdf_output_path = Check_If_Exists(os.path.join(script_dir, os.path.splitext(os.path.basename(file_path))[0]))
        convert(file_path, pdf_output_path)
        logging.info(f"Файл {pdf_output_path} успешно создан")
    except Exception as e: # обработка исключений
        logging.error(f"Ошибка при конвертации файла Word: {e}")
        messagebox.showerror("Ошибка", f"Ошибка при конвертации файла Word: {e}")
        sys.exit(1)


# main функция для выбора файла и запуска конвертации
def main():
    root = Tk()
    root.withdraw()

    logging.info("Выбор файла")

    # выбор файла с помощью диалогового окна
    file_path = filedialog.askopenfilename(
        title = "Выберите файл",
        filetypes = [("All files", "*.*"), ("Excel files", "*.xlsx"), ("Word files", "*.docx")],
        initialdir = os.path.dirname(__file__)
    )

    # обработка исключения, если файл не выбран
    if not file_path:
        logging.error("Файл не выбран")
        messagebox.showerror("Ошибка", "Файл не выбран")
        sys.exit(1)
    else:
        logging.info(f"Выбран файл {file_path}")

    _, file_extension = os.path.splitext(file_path)

    # проверка расширения файла
    if file_extension == '.xlsx':
        logging.info("Началась конвертация файла Excel в PDF")
        xlsx_to_pdf(file_path)
    elif file_extension == '.docx':
        logging.info("Началась конвертация файла Word в PDF")
        convert_docx_to_pdf(file_path)
    else:
        logging.error("Неподдерживаемый формат файла")
        messagebox.showerror("Ошибка", "Неподдерживаемый формат файла")
        sys.exit(1)
    messagebox.showinfo("Успех", "Файл успешно создан!")

# запуск main функции
if __name__ == "__main__":
    main()